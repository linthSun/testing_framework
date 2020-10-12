from openpyxl import *
import os 
# import xlrd

class excel():
    def __init__(self,file):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]

    # 获取表格的总行数和总列数
    @property
    def rows(self):
        rows = self.ws.max_row
        return rows

    @property
    def columns(self):
        columns = self.ws.max_column
        return columns

    # 获取某个单元格的值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 获取某列的所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    # 设置某个单元格的值
    def set_cell_value(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)

    def read_data(self):
        if self.rows > 1 :
             # 获取第一行的内容，列表格式
             keys = self.get_row_value(1)
             listApiData = []
             # 获取每一行的内容，列表格式
             for col in range(2, self.rows+1):
                 values = self.get_row_value(col)
                 # keys，values组合转换为字典
                 api_dict = dict(zip(keys, values))
                 listApiData.append(api_dict)
             return listApiData
        else:
             print("表格没有测试数据!")
             return None


# class ReadExcel():

#      def __init__(self,fileName, SheetName="Sheet1"):
#          self.data = xlrd.open_workbook(fileName)
#          self.table = self.data.sheet_by_name(SheetName)
#            # 获取总行数、总列数
#          self.ncols = self.table.ncols
#          self.nrows = self.table.nrows
#      def read_data(self):
#          if self.nrows > 1:
#              # 获取第一行的内容，列表格式
#              keys = self.table.row_values(0)
#              listApiData = []
#              # 获取每一行的内容，列表格式
#              for col in range(1, self.nrows):
#                  values = self.table.row_values(col)
#                  # keys，values组合转换为字典
#                  api_dict = dict(zip(keys, values))
#                  listApiData.append(api_dict)
#              return listApiData
#          else:
#              print("表格是空数据!")
#              return None

if __name__ == "__main__":
    # test_obj = ReadExcel("test_data.xlsx")
    # print(test_obj.read_data())
    base_dir = os.path.dirname((os.path.dirname(__file__)))
    test_case_file = os.path.join(base_dir,'data','test_data.xlsx')
    test_obj = excel(test_case_file)
    print(test_obj.read_data())